import csv
import os
import sys
import traceback

import pyperclip
from PyQt5.QtWidgets import QDialogButtonBox, QAbstractButton, QMessageBox, QApplication, QStyleFactory, QFileDialog, \
    QAbstractItemView, QTableWidget, QHeaderView, QTableWidgetItem
from openpyxl import load_workbook
from osgeo import ogr

from UI.UILandUseTypeConvert import Ui_Dialog
from PyQt5.QtCore import Qt, QEvent, QThread
from PyQt5 import QtWidgets, QtGui, QtCore

from UICore.DataFactory import workspaceFactory, read_table_header
from UICore.Gv import SplitterState, Dock, DataType
from UICore.common import get_suffix, is_already_opened_in_write_mode
from UICore.log4p import Log
import UI.listview_dialog
from UICore.workerThread import updateAttributeValueWorker

Slot = QtCore.pyqtSlot

log = Log(__name__)


class Ui_Window(QtWidgets.QDialog, Ui_Dialog):
    def __init__(self, parent=None):
        super(Ui_Window, self).__init__(parent=parent)
        self.setupUi(self)

        font = QtGui.QFont()
        font.setFamily("微软雅黑")
        font.setPointSize(9)
        self.buttonBox.setStandardButtons(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self.buttonBox.button(QDialogButtonBox.Ok).setFont(font)
        self.buttonBox.button(QDialogButtonBox.Ok).setText("确定")
        self.buttonBox.button(QDialogButtonBox.Cancel).setFont(font)
        self.buttonBox.button(QDialogButtonBox.Cancel).setText("取消")

        vlayout = QtWidgets.QVBoxLayout(self)
        vlayout.addWidget(self.splitter)
        vlayout.setContentsMargins(0, 0, 10, 10)
        self.splitter.setGeometry(0, 0, self.width(), self.height())

        self.splitter.setOrientation(Qt.Horizontal)
        self.splitter.setProperty("Stretch", SplitterState.collapsed)
        self.splitter.setProperty("Dock", Dock.right)
        self.splitter.setProperty("WidgetToHide", self.txt_log)
        self.splitter.setProperty("ExpandParentForm", True)

        log.setLogViewer(parent=self, logViewer=self.txt_log)
        self.txt_log.setReadOnly(True)

        self.splitter.setSizes([600, self.splitter.width() - 590])
        self.resize(self.splitter.width(), self.splitter.height())

        # self.paras = {}  # 存储参数信息
        # self.selIndex = QModelIndex()
        # self.table_init()

        log.setLogViewer(parent=self, logViewer=self.txt_log)
        self.txt_log.setReadOnly(True)

        self.buttonBox.clicked.connect(self.buttonBox_clicked)
        self.btn_addressLayerFile.clicked.connect(self.btn_addressLayerFile_clicked)
        self.btn_addressConfigFile.clicked.connect(self.btn_addressConfigFile_clicked)
        self.rbtn_file.clicked.connect(self.rbtn_toggled)
        self.rbtn_filedb.clicked.connect(self.rbtn_toggled)
        # self.tableWidget.viewport().installEventFilter(self)
        self.tableWidget.installEventFilter(self)
        # self.splitter.splitterMoved.connect(self.splitterMoved)
        # self.splitter.handle(1).handleClicked.connect(self.handleClicked)

        self.splitter.setupUi()
        self.bInit = True
        self.dialog_width = 0
        self.dialog_height = 0
        # self.rel_tables = []

    def showEvent(self, a0: QtGui.QShowEvent) -> None:
        log.setLogViewer(parent=self, logViewer=self.txt_log)
        if self.bInit:
            self.rbtn_file.click()
            self.bInit = False
            self.dialog_width = self.width()
            self.dialog_height = self.height()

    def threadTerminate(self):
        try:
            if self.thread.isRunning():
                self.thread.terminate()
                self.thread.wait()
                del self.thread
            else:
                self.thread.quit()
                self.thread.wait()
        except:
            return

    def threadStop(self):
        self.thread.quit()


    @Slot()
    def rbtn_toggled(self):
        if not self.bInit:
            self.dialog_init()

    def dialog_init(self):
        self.txt_addressLayerFile.setText("")
        self.txt_addressConfigFile.setText("")
        self.tableWidget.clear()
        self.tableWidget.setRowCount(0)
        self.tableWidget.setColumnCount(0)
        self.resize(self.dialog_width, self.dialog_height)

    @Slot(QAbstractButton)
    def buttonBox_clicked(self, button: QAbstractButton):
        if button == self.buttonBox.button(QDialogButtonBox.Ok):
            DLBM_index = -1
            all_data = []
            layer = None
            header = []

            if self.txt_addressLayerFile.text() == '' or self.txt_addressLayerFile.text() == '':
                return

            self.thread = QThread(self)
            self.updateThread = updateAttributeValueWorker()
            self.updateThread.moveToThread(self.thread)
            self.updateThread.update.connect(self.updateThread.updateAttribute)
            self.updateThread.finished.connect(self.threadStop)

            for icol in range(self.tableWidget.columnCount()):
                header_txt = self.tableWidget.horizontalHeaderItem(icol).text()
                # print(self.tableWidget.horizontalHeaderItem(icol).text())
                if header_txt.upper() == 'DLBM':
                    DLBM_index = icol
                header.append(self.tableWidget.horizontalHeaderItem(icol).text())

            all_data.append(header)

            for irow in range(self.tableWidget.rowCount()):
                row = []
                for icol in range(self.tableWidget.columnCount()):
                    row.append(self.tableWidget.item(irow, icol).text())
                all_data.append(row)

            rel_tables = self.generate_config_rel(DLBM_index, all_data)

            fileName = self.txt_addressLayerFile.text()
            fileName_arr = fileName.split(os.sep)
            in_path = os.sep.join(fileName_arr[:-1])
            layer_name = fileName_arr[-1]

            file_type = None
            if self.rbtn_file.isChecked():
                wks = workspaceFactory().get_factory(DataType.shapefile)
                datasource = wks.openFromFile(fileName, 1)
                layer = datasource.GetLayer(0)
                file_type = DataType.shapefile
                in_path = fileName
            elif self.rbtn_filedb.isChecked():
                wks = workspaceFactory().get_factory(DataType.fileGDB)
                datasource = wks.openFromFile(in_path, 1)
                layer = datasource.GetLayerByName(layer_name)
                file_type = DataType.fileGDB

            datasource = None

            if layer is not None:
                self.thread.start()
                self.updateThread.update.emit(file_type, in_path, layer_name, header, rel_tables)
            else:
                log.error("无法读取矢量数据！请检查路径和数据完整性", dialog=True)
            print("OK")
        elif button == self.buttonBox.button(QDialogButtonBox.Cancel):
            self.threadTerminate()
            self.close()

    @Slot()
    def btn_addressLayerFile_clicked(self):
        # _f_dlg = FileDialog()
        # _f_dlg.exec_()
        datasource = None
        layer = None
        try:
            if self.rbtn_file.isChecked():
                fileName, fileType = QtWidgets.QFileDialog.getOpenFileName(
                    self, "选择待转换矢量图层文件", os.getcwd(),
                    "ESRI Shapefile(*.shp)")

                if len(fileName) == 0:
                    return

                if is_already_opened_in_write_mode(fileName):
                    log.error("矢量图层正在被占用，请先关闭其他应用程序！", dialog=True)
                    return

                fileType = get_suffix(fileName)

                if fileType == DataType.shapefile:
                    wks = workspaceFactory().get_factory(DataType.shapefile)
                    datasource = wks.openFromFile(fileName)
                    self.txt_addressLayerFile.setText(fileName)
                else:
                    log.error("不识别的图形文件格式！", dialog=True)

                if datasource is not None:
                    layer = datasource.GetLayer(0)
                else:
                    log.error("无法读取shp文件！{}".format(fileName), dialog=True)
            elif self.rbtn_filedb.isChecked():
                fileName = QtWidgets.QFileDialog.getExistingDirectory(self, "选择需要转换的GDB数据库",
                                                                      os.getcwd(), QFileDialog.ShowDirsOnly)

                if is_already_opened_in_write_mode(fileName):
                    log.error("矢量图层正在被占用，请先关闭其他应用程序！", dialog=True)
                    return

                wks = workspaceFactory().get_factory(DataType.fileGDB)
                datasource = wks.openFromFile(fileName)

                selected_name = None
                if datasource is not None:
                    lst_names = wks.getLayerNames()
                    if len(lst_names) > 1:
                        selected_name = nameListDialog().openListDialog("请选择要转换的图层", lst_names)
                    elif len(lst_names) == 1:
                        selected_name = [lst_names[0]]

                    layerName = selected_name[0]
                    self.txt_addressLayerFile.setText(os.path.join(fileName, layerName))
                    layer = datasource.GetLayerByName(layerName)
                else:
                    log.error("无法读取文件数据库！{}".format(fileName), dialog=True)

            if layer is not None:
                self.check_field(layer)
        except:
            log.error("无法读取矢量数据图层.\n" + traceback.format_exc())
        finally:
            datasource = None
            layer = None

    @Slot()
    def btn_addressConfigFile_clicked(self):
        fileName, types = QFileDialog.getOpenFileName(
            self, "选择土地类型转换规则表文件", os.getcwd(),
            "表格文件(*.csv *.xlsx);;csv文件(*.csv);;excel文件(*.xlsx)")

        if len(fileName) == 0:
            return

        fileType = get_suffix(fileName)

        if fileType != DataType.xlsx and fileType != DataType.csv:
            log.error("不识别的图形文件格式！", dialog=True)
            return None

        header, DLBM_values, all_data = self.read_config_table(fileName, fileType)

        DLBM_index = self.check_header(header)
        if DLBM_index > -1:
            if self.check_field_DLBM(DLBM_values):
                self.add_all_data_to_tablewidget(DLBM_index, all_data)
                self.txt_addressConfigFile.setText(fileName)
                # self.rel_tables = self.generate_config_rel(DLBM_index, all_data)

    def eventFilter(self, source: 'QObject', event: 'QEvent') -> bool:
        if source is self.tableWidget and event.type() == QEvent.KeyPress:
            if event.modifiers() == Qt.ControlModifier:
                if event.key() == Qt.Key_V:
                    file = pyperclip.paste()
                    header, DLBM_values, all_data = self.read_config_table(file, fileType=DataType.memory)

                    DLBM_index = self.check_header(header)
                    if DLBM_index > -1:
                        if self.check_field_DLBM(DLBM_values):
                            self.add_all_data_to_tablewidget(DLBM_index, all_data)
                            # self.rel_tables = self.generate_config_rel(DLBM_index, all_data)
        # if source is self.tableWidget.viewport():
        #     print(event.type())
        # # if source is self.tableWidget.viewport() and event.type() == QEvent.MouseButtonPress:
        # #     print(event.type())
        return super().eventFilter(source, event)

    #  在表格控件中显示读取的规则配置表数据
    def add_all_data_to_tablewidget(self, DLBM_index, all_data):
        col_num = len(all_data[0])
        row_num = len(all_data)

        self.tableWidget.setColumnCount(col_num)
        self.tableWidget.setRowCount(row_num - 1)
        self.tableWidget.setHorizontalHeaderLabels(all_data[0])
        self.tableWidget.horizontalHeader().setSectionsMovable(False)  # 表头顺序不允许调整

        # 按行加载数据
        irow = 0
        for row_value in all_data:
            if irow == 0:
                irow += 1
                continue

            for icol in range(len(row_value)):
                newItem = QTableWidgetItem(row_value[icol])
                if icol == DLBM_index:
                    newItem.setFlags(QtCore.Qt.ItemIsEnabled)  # DLBM字段设置为不可编辑
                self.tableWidget.setItem(irow - 1, icol, newItem)

            irow += 1

        # 首先根据resizetocontents自动计算合适的列宽
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

        # 首先根据resizetocontents自动计算合适的列宽，然后自动调整窗口大小，最后将列宽设置为可调整
        resize_width = 0
        for i in range(self.tableWidget.columnCount()):
            resize_width = resize_width + self.tableWidget.columnWidth(i)

        if self.splitter.splitterState == SplitterState.collapsed:
            init_width = self.tableWidget.width()
            resize_width = self.width() + (resize_width - init_width)
            self.resize(resize_width + 40, self.height())
        else:
            init_width = self.splitter.widget(0).width()
            resize_width2 = self.splitter.width() + (resize_width - init_width)
            self.resize(resize_width2, self.height())
            self.splitter.setSizes([resize_width, self.splitter.width() - resize_width - 30])

        self.tableWidget.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)  # 行高固定
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)

    def read_config_table(self, file, fileType):
            all_data = []
            not_none_index = []
            DLBM_values = []
            DLBM_index = -1
            header = None

            if fileType == DataType.xlsx:
                wb = load_workbook(file, read_only=True)

                selected_sheet = []

                lst_names = wb.sheetnames
                if len(lst_names) > 1:
                    selected_sheet = nameListDialog().openListDialog(
                        "请选择工作表(sheet)", lst_names, QAbstractItemView.SingleSelection)
                    selected_sheet = selected_sheet[0]
                elif len(lst_names) == 1:
                    selected_sheet = lst_names[0]

                all_values = list(wb[selected_sheet].values)

                header, bheader = read_table_header(file, fileType, supplyment=False, sheet=selected_sheet)

                # 表头非空的列才计入规则表
                for i in range(len(header)):
                    if header[i] != 'None':
                        not_none_index.append(i)
                    if header[i].upper() == 'DLBM':
                        DLBM_index = i

                for row_values in all_values:
                    row_values_filter = []
                    for i in range(len(row_values)):
                        if i in not_none_index:
                            row_values_filter.append(row_values[i])
                        if i == DLBM_index:
                            DLBM_values.append(row_values[i])

                    all_data.append(row_values_filter)

                wb.close()

            elif fileType == DataType.csv:
                header, encoding, bheader = read_table_header(file, fileType, supplyment=False, sheet=None)
                for i in range(len(header)):
                    if header[i] != '':
                        not_none_index.append(i)
                    if header[i].upper() == 'DLBM':
                        DLBM_index = i

                with open(file, 'r', newline='', encoding=encoding) as f:
                    reader = csv.reader(f)
                    for row in reader:
                        row_values_filter = []
                        for i in range(len(row)):
                            if i in not_none_index:
                                row_values_filter.append(row[i])
                            if i == DLBM_index:
                                DLBM_values.append(row[i])

                        all_data.append(row_values_filter)

            elif fileType == DataType.memory:
                header, bheader = read_table_header(file, fileType, supplyment=False, sheet=None)
                for i in range(len(header)):
                    if header[i] != '':
                        not_none_index.append(i)
                    if header[i].upper() == 'DLBM':
                        DLBM_index = i

                for line in file.splitlines():
                    row = line.split('\t')
                    row_values_filter = []
                    for i in range(len(row)):
                        if i in not_none_index:
                            row_values_filter.append(row[i])
                        if i == DLBM_index:
                            DLBM_values.append(row[i])

                    all_data.append(row_values_filter)

            return header, DLBM_values, all_data

    def check_field(self, layer):
        bexist = False
        layerDefn = layer.GetLayerDefn()
        for i in range(layerDefn.GetFieldCount()):
            fieldName = layerDefn.GetFieldDefn(i).GetName()
            if fieldName.upper() == "DLBM":
                bexist = True
                break
        if not bexist:
            log.error('矢量图层数据缺失必要字段"DLMB"，请补全！', dialog=True)
            return False

        return True

    # 检查表头和数据的合法性
    def check_header(self, header):
        DLBM_index = -1
        not_none = []
        for i in header:
            if i != '':
                not_none.append(i)

        for i in range(len(not_none)):
            if not_none[i].upper() == 'DLBM':
                DLBM_index = i
                break

        if DLBM_index == -1:
            log.error("规则表不存在必要字段DLBM，请检查！", dialog=True)
        if len(not_none) <= DLBM_index:
            log.error("DLBM列右边缺少需要匹配的列，请检查！", dialog=True)

        return DLBM_index

    def check_field_DLBM(self, DLBM):
        set_headers = set(DLBM)
        if len(set_headers) != len(DLBM):
            log.error("规则表DLBM字段不允许出现重复值，请检查！", dialog=True)
            return False
        return True

    # 生成对应规则字典, DLBM列是唯一KEY,右边的列都是VALUE
    def generate_config_rel(self, DLBM_INDEX, all_data):
        header = all_data[0]
        rel_tables = []  # 存储所有的规则关系字典

        for icol in range(DLBM_INDEX + 1, len(header)):
            iFirst = 0
            rel = {}

            for row_value in all_data:
                if iFirst == 0:
                    iFirst += 1
                    continue
                rel[row_value[DLBM_INDEX]] = row_value[icol]

            rel_tables.append(rel)

        return rel_tables


class nameListDialog(QtWidgets.QDialog, UI.listview_dialog.Ui_Dialog):
    def __init__(self):
        super(nameListDialog, self).__init__()
        self.setupUi(self)
        self.pushButton.clicked.connect(self.pushButton_clicked)
        self.select_names = []

    def openListDialog(self, title, lst_names, selectMode=QAbstractItemView.SingleSelection):
        self.lv_name.setSelectionMode(selectMode)
        self.setWindowTitle(title)

        for name in lst_names:
            self.lv_name.addItem(name)

        result = self.exec_()

        if result == 1:
            return self.select_names

    def pushButton_clicked(self):
        sel_items = self.lv_name.selectedItems()
        for item in sel_items:
            self.select_names.append(item.text())
        self.done(1)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    style = QStyleFactory.create("windows")
    app.setStyle(style)
    window = Ui_Window()
    window.setWindowFlags(Qt.Window)
    window.show()
    sys.exit(app.exec_())